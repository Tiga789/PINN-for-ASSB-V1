#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Prepare cycle-level capacity/SOH targets for ModelFin_108 capacity-physics training.

This script is intentionally independent from the existing PINN training code so that
training and evaluation can share exactly the same capacity target CSV.

Primary supported sources
-------------------------
1) ZHB_ASSB_NCM811.xlsx, sheet "step", column "放电容量(Ah)".
2) record_extracted.csv, time-record table with discharge-capacity column.

Output columns
--------------
cycle_id, Q_dis_Ah, Q_dis_mAh, Q_ref_Ah, SOH, complete_cycle, train_mask,
V_min, V_end, I_dis_abs_A, discharge_step_count, source_path, source_sheet.

Notes
-----
- Cycle 1-4 are excluded by default through --cycle_from 5.
- Partial cycles are retained in the CSV but can be excluded from training.
- The default Q_ref is max Q_dis over complete cycles in cycle 5-20.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np

try:
    import pandas as pd
except Exception as exc:  # pragma: no cover
    pd = None


# -----------------------------------------------------------------------------
# Column-name helpers
# -----------------------------------------------------------------------------


def _norm_name(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = text.replace("_", "").replace("-", "")
    return text


def _find_col(headers: Sequence[Any], candidates: Sequence[str], *, required: bool = True) -> Optional[str]:
    norm_headers = {_norm_name(h): str(h).strip() for h in headers if h is not None and str(h).strip()}
    norm_candidates = [_norm_name(c) for c in candidates]

    # Exact normalized match first.
    for cand in norm_candidates:
        if cand in norm_headers:
            return norm_headers[cand]

    # Substring match second, favor longer headers.
    for cand in norm_candidates:
        for nh, original in sorted(norm_headers.items(), key=lambda kv: -len(kv[0])):
            if cand and cand in nh:
                return original

    if required:
        raise KeyError(f"Could not find required column. Candidates={candidates}; headers={list(headers)}")
    return None


CYCLE_COLS = ["循环号", "循环", "cycle_id", "cycle", "cycleindex", "循环序号"]
STEP_TYPE_COLS = ["工步类型", "步骤类型", "step_type", "steptype", "step", "工步"]
STEP_ID_COLS = ["工步号", "工步序号", "step_id", "stepid", "step index"]
Q_DIS_COLS = ["放电容量(Ah)", "放电容量", "discharge capacity(ah)", "discharge_capacity(ah)", "dchg cap(ah)", "dchg. cap.(ah)"]
Q_CHG_COLS = ["充电容量(Ah)", "充电容量", "charge capacity(ah)", "charge_capacity(ah)", "chg cap(ah)"]
Q_ANY_COLS = ["容量(Ah)", "capacity(ah)", "capacity"]
VOLTAGE_COLS = ["电压(V)", "电压", "voltage(v)", "voltage", "结束电压(V)", "end voltage(v)"]
V_END_COLS = ["结束电压(V)", "end voltage(v)", "end_voltage", "voltage_end"]
V_START_COLS = ["起始电压(V)", "start voltage(v)", "start_voltage", "voltage_start"]
CURRENT_COLS = ["电流(A)", "电流", "current(a)", "current"]
TIME_COLS = ["工步时间", "step time", "steptime", "time", "时间", "test time(s)"]


def _looks_like_discharge(series: Iterable[Any]) -> np.ndarray:
    vals = ["" if x is None else str(x).strip().lower() for x in series]
    return np.array([("放电" in v) or ("discharge" in v) or ("dchg" in v) for v in vals], dtype=bool)


def _to_float_array(values: Iterable[Any], default: float = np.nan) -> np.ndarray:
    out: List[float] = []
    for v in values:
        if v is None or v == "":
            out.append(default)
            continue
        try:
            out.append(float(v))
        except Exception:
            text = str(v).replace(",", "").strip()
            try:
                out.append(float(text))
            except Exception:
                out.append(default)
    return np.asarray(out, dtype=float)


def _parse_step_time_seconds(value: Any) -> float:
    """Parse HH:MM:SS-like step time to seconds. Returns nan if unavailable."""
    if value is None or value == "":
        return float("nan")
    if isinstance(value, (int, float)):
        # Excel time may be stored as fraction of a day; raw seconds may also appear.
        v = float(value)
        if 0.0 <= v <= 10.0:
            return v * 86400.0
        return v
    text = str(value).strip()
    m = re.match(r"^(?:(\d+)[dD]\s*)?(\d{1,3}):(\d{2}):(\d{2}(?:\.\d+)?)$", text)
    if m:
        days = int(m.group(1) or 0)
        h = int(m.group(2))
        mi = int(m.group(3))
        s = float(m.group(4))
        return days * 86400.0 + h * 3600.0 + mi * 60.0 + s
    # Some BTS exports use "12:34".
    m = re.match(r"^(\d{1,3}):(\d{2})$", text)
    if m:
        return int(m.group(1)) * 60.0 + int(m.group(2))
    return float("nan")


# -----------------------------------------------------------------------------
# Readers
# -----------------------------------------------------------------------------


def _read_csv_table(path: Path, encoding: Optional[str] = None) -> "pd.DataFrame":
    if pd is None:
        raise RuntimeError("pandas is required to read CSV files in this script.")
    encodings = [encoding] if encoding else ["utf-8-sig", "utf-8", "gbk", "gb18030"]
    last_exc: Optional[Exception] = None
    for enc in encodings:
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception as exc:
            last_exc = exc
    raise RuntimeError(f"Failed to read CSV {path}: {last_exc}")


def _select_sheet_name(workbook: Any, sheet_hint: Optional[str]) -> str:
    names = list(workbook.sheetnames)
    if sheet_hint:
        if sheet_hint in names:
            return sheet_hint
        low = sheet_hint.strip().lower()
        for name in names:
            if low == name.lower():
                return name
        for name in names:
            if low in name.lower():
                return name
        raise KeyError(f"Excel sheet {sheet_hint!r} not found. Available: {names}")
    for name in names:
        if "step" in name.lower() or "工步" in name:
            return name
    return names[0]


def _read_xlsx_sheet(path: Path, sheet_name: Optional[str]) -> Tuple["pd.DataFrame", str]:
    """Read an Excel sheet robustly, including sheets with broken dimension metadata.

    The uploaded ZHB workbook has useful sheet names (unit/test/cycle/step/record/...),
    and some sheets can report an incorrect A1 dimension. We therefore reset dimensions
    before iterating rows.
    """
    if pd is None:
        raise RuntimeError("pandas is required to build DataFrame outputs.")
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read .xlsx capacity sources.") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    selected = _select_sheet_name(wb, sheet_name)
    ws = wb[selected]
    try:
        ws.reset_dimensions()
    except Exception:
        pass

    rows_iter = ws.iter_rows(values_only=True)
    rows: List[Tuple[Any, ...]] = []
    for row in rows_iter:
        # Trim trailing empty cells but keep internal Nones.
        r = tuple(row)
        while r and r[-1] is None:
            r = r[:-1]
        if r:
            rows.append(r)

    if not rows:
        raise RuntimeError(f"Excel sheet {selected!r} is empty.")

    # Header row is usually the first row; scan anyway.
    best_idx = 0
    best_score = -1
    for i, row in enumerate(rows[:30]):
        norm = [_norm_name(x) for x in row]
        score = 0
        for cands in [CYCLE_COLS, STEP_TYPE_COLS, Q_DIS_COLS, VOLTAGE_COLS, CURRENT_COLS, V_END_COLS]:
            if any(any(_norm_name(c) in h for h in norm) for c in cands):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = i
    headers = [str(h).strip() if h is not None else f"col_{j+1}" for j, h in enumerate(rows[best_idx])]
    data_rows = rows[best_idx + 1 :]
    width = len(headers)
    matrix = [list(r[:width]) + [None] * max(0, width - len(r)) for r in data_rows]
    df = pd.DataFrame(matrix, columns=headers)
    return df, selected


def read_source_table(source_path: Path, *, excel_sheet: Optional[str] = "step", csv_encoding: Optional[str] = None) -> Tuple["pd.DataFrame", Optional[str]]:
    ext = source_path.suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_xlsx_sheet(source_path, excel_sheet)
    if ext in {".csv", ".txt"}:
        return _read_csv_table(source_path, csv_encoding), None
    raise ValueError(f"Unsupported source type {ext}; use .xlsx or .csv")


# -----------------------------------------------------------------------------
# Aggregation
# -----------------------------------------------------------------------------


@dataclass
class CapacityBuildOptions:
    cycle_from: int = 5
    cycle_to: Optional[int] = None
    qref_cycle_from: int = 5
    qref_cycle_to: int = 20
    complete_v_max: float = 2.20
    min_q_ah: float = 2.5e-4
    exclude_incomplete: bool = True
    q_ref_ah: Optional[float] = None


def build_capacity_targets(df: "pd.DataFrame", opts: CapacityBuildOptions, *, source_path: str, source_sheet: Optional[str]) -> Tuple["pd.DataFrame", Dict[str, Any]]:
    if pd is None:
        raise RuntimeError("pandas is required.")

    headers = list(df.columns)
    cycle_col = _find_col(headers, CYCLE_COLS, required=True)
    step_type_col = _find_col(headers, STEP_TYPE_COLS, required=False)
    q_dis_col = _find_col(headers, Q_DIS_COLS, required=False)
    q_any_col = _find_col(headers, Q_ANY_COLS, required=False)
    q_chg_col = _find_col(headers, Q_CHG_COLS, required=False)
    v_col = _find_col(headers, VOLTAGE_COLS, required=False)
    v_end_col = _find_col(headers, V_END_COLS, required=False)
    v_start_col = _find_col(headers, V_START_COLS, required=False)
    i_col = _find_col(headers, CURRENT_COLS, required=False)
    time_col = _find_col(headers, TIME_COLS, required=False)
    step_id_col = _find_col(headers, STEP_ID_COLS, required=False)

    work = df.copy()
    work["__cycle_id"] = _to_float_array(work[cycle_col]).astype(float)
    work = work[np.isfinite(work["__cycle_id"])]
    work["__cycle_id"] = work["__cycle_id"].astype(int)
    work = work[work["__cycle_id"] >= int(opts.cycle_from)]
    if opts.cycle_to is not None:
        work = work[work["__cycle_id"] <= int(opts.cycle_to)]

    if q_dis_col is not None:
        work["__q_dis_ah"] = _to_float_array(work[q_dis_col])
    elif q_any_col is not None:
        work["__q_dis_ah"] = _to_float_array(work[q_any_col])
    else:
        raise KeyError("Cannot find discharge capacity column. Expected '放电容量(Ah)' or equivalent.")

    if q_chg_col is not None:
        work["__q_chg_ah"] = _to_float_array(work[q_chg_col])
    else:
        work["__q_chg_ah"] = np.nan

    if step_type_col is not None:
        dis_mask = _looks_like_discharge(work[step_type_col].values)
    else:
        dis_mask = np.isfinite(work["__q_dis_ah"].values) & (work["__q_dis_ah"].values > 0)
        if i_col is not None:
            current = _to_float_array(work[i_col])
            # In this project convention, negative current is discharge for record data.
            # Step tables may not contain current; do not require it.
            dis_mask = dis_mask & ((current < 0) | ~np.isfinite(current))

    work = work[dis_mask].copy()
    if len(work) == 0:
        raise RuntimeError("No discharge rows found after filtering. Check step type and capacity columns.")

    if v_col is not None:
        work["__v"] = _to_float_array(work[v_col])
    elif v_end_col is not None:
        work["__v"] = _to_float_array(work[v_end_col])
    else:
        work["__v"] = np.nan

    if v_end_col is not None:
        work["__v_end"] = _to_float_array(work[v_end_col])
    else:
        work["__v_end"] = work["__v"].values

    if v_start_col is not None:
        work["__v_start"] = _to_float_array(work[v_start_col])
    else:
        work["__v_start"] = np.nan

    if i_col is not None:
        work["__i_abs"] = np.abs(_to_float_array(work[i_col]))
    else:
        work["__i_abs"] = np.nan

    if time_col is not None:
        work["__step_time_s"] = [_parse_step_time_seconds(x) for x in work[time_col].values]
    else:
        work["__step_time_s"] = np.nan

    if step_id_col is not None:
        work["__step_id"] = _to_float_array(work[step_id_col])
    else:
        work["__step_id"] = np.nan

    rows: List[Dict[str, Any]] = []
    for cid, g in work.groupby("__cycle_id", sort=True):
        q_dis = float(np.nanmax(g["__q_dis_ah"].values))
        q_chg_vals = g["__q_chg_ah"].values
        q_chg = float(np.nanmax(q_chg_vals)) if np.any(np.isfinite(q_chg_vals)) else float("nan")
        v_vals = g["__v"].values
        v_min = float(np.nanmin(v_vals)) if np.any(np.isfinite(v_vals)) else float("nan")
        v_end_vals = g["__v_end"].values
        v_end = float(v_end_vals[np.where(np.isfinite(v_end_vals))[0][-1]]) if np.any(np.isfinite(v_end_vals)) else float("nan")
        i_vals = g["__i_abs"].values
        i_abs = float(np.nanmax(i_vals)) if np.any(np.isfinite(i_vals)) else float("nan")
        t_vals = g["__step_time_s"].values
        step_time_s = float(np.nansum(t_vals)) if np.any(np.isfinite(t_vals)) else float("nan")
        rows.append(
            {
                "cycle_id": int(cid),
                "Q_dis_Ah": q_dis,
                "Q_dis_mAh": q_dis * 1000.0,
                "Q_chg_Ah": q_chg,
                "V_min": v_min,
                "V_end": v_end,
                "I_dis_abs_A": i_abs,
                "discharge_step_time_s": step_time_s,
                "discharge_step_count": int(len(g)),
                "first_step_id": float(np.nanmin(g["__step_id"].values)) if np.any(np.isfinite(g["__step_id"].values)) else float("nan"),
                "last_step_id": float(np.nanmax(g["__step_id"].values)) if np.any(np.isfinite(g["__step_id"].values)) else float("nan"),
            }
        )

    out = pd.DataFrame(rows).sort_values("cycle_id").reset_index(drop=True)
    has_voltage = np.isfinite(out["V_min"].values).any()
    if has_voltage:
        complete = (out["V_min"].values <= float(opts.complete_v_max)) & (out["Q_dis_Ah"].values >= float(opts.min_q_ah))
    else:
        complete = out["Q_dis_Ah"].values >= float(opts.min_q_ah)
    out["complete_cycle"] = complete.astype(bool)
    out["partial_or_incomplete"] = ~out["complete_cycle"].values

    if opts.q_ref_ah is not None and opts.q_ref_ah > 0:
        q_ref = float(opts.q_ref_ah)
        q_ref_mode = "manual"
    else:
        qmask = (
            (out["cycle_id"] >= int(opts.qref_cycle_from))
            & (out["cycle_id"] <= int(opts.qref_cycle_to))
            & out["complete_cycle"].astype(bool)
        )
        if not qmask.any():
            qmask = out["complete_cycle"].astype(bool)
        if not qmask.any():
            qmask = out["Q_dis_Ah"] > 0
        q_ref = float(out.loc[qmask, "Q_dis_Ah"].max())
        q_ref_mode = f"max_complete_cycle_{opts.qref_cycle_from}_{opts.qref_cycle_to}"

    if not math.isfinite(q_ref) or q_ref <= 0:
        raise RuntimeError(f"Invalid Q_ref_Ah={q_ref}. Check capacity target table.")

    out["Q_ref_Ah"] = q_ref
    out["Q_ref_mAh"] = q_ref * 1000.0
    out["SOH"] = out["Q_dis_Ah"] / q_ref
    out["SOH_clipped"] = out["SOH"].clip(lower=0.0, upper=1.05)
    out["train_mask"] = out["complete_cycle"].astype(bool) if opts.exclude_incomplete else True
    out["source_path"] = source_path
    out["source_sheet"] = source_sheet or ""

    # A few protocol-only feature hints for downstream capacity head construction.
    cid = out["cycle_id"].astype(float).values
    denom = max(float(cid.max() - cid.min()), 1.0)
    out["cycle_norm"] = (cid - cid.min()) / denom
    dcycle = np.diff(cid, prepend=cid[0])
    if len(dcycle) > 1 and dcycle[0] == 0:
        dcycle[0] = dcycle[1]
    out["d_cycle"] = dcycle
    out["d_tau"] = np.maximum(dcycle / max(float(np.sum(dcycle)), 1.0), 1.0e-8)
    if np.isfinite(out["discharge_step_time_s"].values).any() and np.isfinite(out["I_dis_abs_A"].values).any():
        out["estimated_discharge_throughput_Ah"] = out["I_dis_abs_A"].fillna(0.0) * out["discharge_step_time_s"].fillna(0.0) / 3600.0
    else:
        out["estimated_discharge_throughput_Ah"] = np.nan

    meta = {
        "source_path": source_path,
        "source_sheet": source_sheet,
        "cycle_from": opts.cycle_from,
        "cycle_to": opts.cycle_to,
        "q_ref_Ah": q_ref,
        "q_ref_mAh": q_ref * 1000.0,
        "q_ref_mode": q_ref_mode,
        "qref_cycle_from": opts.qref_cycle_from,
        "qref_cycle_to": opts.qref_cycle_to,
        "complete_v_max": opts.complete_v_max,
        "min_q_ah": opts.min_q_ah,
        "exclude_incomplete": opts.exclude_incomplete,
        "n_cycles_total": int(len(out)),
        "n_train_cycles": int(out["train_mask"].astype(bool).sum()),
        "n_incomplete_cycles": int((~out["complete_cycle"].astype(bool)).sum()),
        "cycle_min": int(out["cycle_id"].min()) if len(out) else None,
        "cycle_max": int(out["cycle_id"].max()) if len(out) else None,
        "Q_dis_min_mAh": float(out["Q_dis_mAh"].min()) if len(out) else None,
        "Q_dis_max_mAh": float(out["Q_dis_mAh"].max()) if len(out) else None,
        "SOH_min": float(out["SOH"].min()) if len(out) else None,
        "SOH_max": float(out["SOH"].max()) if len(out) else None,
        "columns_detected": {
            "cycle_col": cycle_col,
            "step_type_col": step_type_col,
            "q_dis_col": q_dis_col,
            "q_chg_col": q_chg_col,
            "voltage_col": v_col,
            "v_end_col": v_end_col,
            "current_col": i_col,
            "time_col": time_col,
            "step_id_col": step_id_col,
        },
    }
    return out, meta


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build cycle-level ASSB capacity/SOH targets from ZHB step Excel or record CSV.")
    p.add_argument("--source_path", default=None, help="Input .xlsx or .csv. Preferred: ZHB_ASSB_NCM811.xlsx.")
    p.add_argument("--record_csv", default=None, help="Backward-compatible alias for --source_path when using record_extracted.csv.")
    p.add_argument("--excel_sheet", default="step", help="Excel sheet name; default uses the 'step' sheet.")
    p.add_argument("--csv_encoding", default=None, help="Optional CSV encoding override.")
    p.add_argument("--cycle_from", type=int, default=5)
    p.add_argument("--cycle_to", type=int, default=522)
    p.add_argument("--qref_cycle_from", type=int, default=5)
    p.add_argument("--qref_cycle_to", type=int, default=20)
    p.add_argument("--q_ref_ah", type=float, default=None, help="Manual Q_ref in Ah. If omitted, max complete cycle in qref range is used.")
    p.add_argument("--complete_v_max", type=float, default=2.20, help="Cycle is complete if V_min <= this threshold, when voltage is available.")
    p.add_argument("--min_q_ah", type=float, default=2.5e-4, help="Minimum discharge capacity to treat a cycle as complete.")
    p.add_argument("--exclude_incomplete", action="store_true", help="Set train_mask=False for incomplete cycles.")
    p.add_argument("--include_incomplete", action="store_true", help="Override --exclude_incomplete.")
    p.add_argument("--output_csv", required=True)
    p.add_argument("--output_json", required=True)
    return p.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    source = args.source_path or args.record_csv
    if not source:
        raise SystemExit("Provide --source_path ZHB_ASSB_NCM811.xlsx or --record_csv record_extracted.csv")
    source_path = Path(source)
    if not source_path.exists():
        raise SystemExit(f"Input source does not exist: {source_path}")

    df, sheet = read_source_table(source_path, excel_sheet=args.excel_sheet, csv_encoding=args.csv_encoding)
    opts = CapacityBuildOptions(
        cycle_from=args.cycle_from,
        cycle_to=args.cycle_to,
        qref_cycle_from=args.qref_cycle_from,
        qref_cycle_to=args.qref_cycle_to,
        complete_v_max=args.complete_v_max,
        min_q_ah=args.min_q_ah,
        exclude_incomplete=(args.exclude_incomplete and not args.include_incomplete),
        q_ref_ah=args.q_ref_ah,
    )
    targets, meta = build_capacity_targets(df, opts, source_path=str(source_path), source_sheet=sheet)

    out_csv = Path(args.output_csv)
    out_json = Path(args.output_json)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    targets.to_csv(out_csv, index=False, encoding="utf-8-sig")
    out_json.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    print("[ASSB capacity targets] OK")
    print(f"  source      : {source_path}")
    print(f"  sheet       : {sheet or '<csv>'}")
    print(f"  output_csv  : {out_csv}")
    print(f"  output_json : {out_json}")
    print(f"  cycles      : {meta['cycle_min']}..{meta['cycle_max']}  n={meta['n_cycles_total']}  train={meta['n_train_cycles']}")
    print(f"  Q_ref       : {meta['q_ref_mAh']:.6f} mAh ({meta['q_ref_mode']})")
    print(f"  SOH range   : {meta['SOH_min']:.6f}..{meta['SOH_max']:.6f}")
    if meta["n_incomplete_cycles"]:
        bad = targets.loc[~targets["complete_cycle"].astype(bool), ["cycle_id", "Q_dis_mAh", "V_min", "V_end"]].tail(10)
        print("  incomplete cycles tail:")
        print(bad.to_string(index=False))
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
